from faker import Faker
from openpyxl import Workbook

wb =Workbook()
ws = wb.active
obj = Faker(['hi_IN' ,'en_US'])

print(obj.name())
print(obj.address())
print(obj.email())
print(obj.text())
print(obj.country())

for i in range (1,11):
    for j in range (1,6):
        ws.cell(row=i,column=1).value = obj.name()
        ws.cell(row=i, column=2).value = obj.address()
        ws.cell(row=i, column=3).value = obj.email()
        ws.cell(row=i, column=4).value = obj.text()
        ws.cell(row=i, column=5).value = obj.country()
wb.save("testdata.xlsx")



