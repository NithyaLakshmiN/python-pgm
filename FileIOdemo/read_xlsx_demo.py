from openpyxl import Workbook,load_workbook

wb =load_workbook(filename ="testdata1.xlsx")
ws = wb['Sheet']

row_count = ws.max_row
col_count = ws.max_column


print(ws.max_row)
print(ws.max_column)

for i in range (1,row_count+1):
    for j in range (1,col_count+1):
        print(ws.cell(row=i,column=j).value , end ='')

    print('\n')


